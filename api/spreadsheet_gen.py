import openpyxl
from openpyxl.styles import PatternFill

def gen_spreadsheet(book_info):
    workbook = openpyxl.load_workbook('template.xlsx')
    sheet = workbook.active
    length = len(book_info)
    green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for i in range(length):
        cell = lambda l: l + str(i+4)
        row = [1, 0, 0, book_info[i][1], None, book_info[i][0], "ISBN", "New"]

        for col, value in enumerate(row, start=1):
            sheet.cell(row=i+4, column=col).value = value

        sheet[cell("C")].fill = green
        sheet[cell("U")] = "Amazon_NA"
        sheet[cell("AB")] = "NO"
        
        for col in range(42, 47):
            sheet.cell(row=i+4, column=col).value = "not_applicable"

    workbook.save('test.xlsx')

gen_spreadsheet([{"price": 27, "isbn": 10}, {"price": 44, "isbn": 9}, {"price": 25, "isbn": 11}], [])