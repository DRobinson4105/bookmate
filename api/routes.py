from pyzbar.pyzbar import decode
from PIL import Image, ImageDraw, ImageFont
from ultralytics import YOLO
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from waitress import serve
from dotenv import load_dotenv
import torch
from torch.utils.data import DataLoader
import os
import io
import base64
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import easyocr
import openpyxl
from openpyxl.styles import PatternFill
import sys

# move to training directory
sys.path.append('../training/prices')
from price_predictor import Model, BookDataset
sys.path.append('../../api')

reader = easyocr.Reader(['en'])

condition_stoi = {'New': 1, 'Like New': 0.85, 'Very Good': 0.7, 'Good': 0.6, 'Acceptable': 0.4}

app = Flask(__name__)
CORS(app)

load_dotenv()

# get api directory
dir = os.path.dirname(os.path.abspath(__file__))

# ISBN Retrieval

barcode_model = YOLO(os.path.join(dir, "barcodeDetection.pt"))
barcode_model.fuse()

@app.route("/api/getISBNs", methods=["POST"])
def get_isbns():
    file = request.files.get('file')
    image_stream = io.BytesIO(file.read())
    image = Image.open(image_stream)
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype("arial.ttf", 72)
    results = barcode_model(image, save=True)

    result = []

    # get bounding box coordinates
    boxes = [box.cpu().numpy().tolist() for box in results[0].obb.xyxyxyxy]
    boxes = [[[int(num) for num in pair] for pair in box] for box in boxes]

    for box in boxes:
        left = 1e9
        top = 1e9
        right = 0
        bottom = 0

        # widen bounding box
        for pair in box:
            left = min(left, pair[0])
            top = min(top, pair[1])
            right = max(right, pair[0])
            bottom = max(bottom, pair[1])

        left, top, right, bottom = left-50, top-50, right+50, bottom+50

        # get isbn from cropped image
        decoded_objects = decode(image.crop((left, top, right, bottom)))

        if len(decoded_objects) == 0: continue
        
        # draw bounding box on image
        draw.rectangle((left, top, right, bottom), outline="black", fill=None, width=5)

        # display isbn on image
        isbn = decoded_objects[0].data.decode("utf-8")
        position = (left, top - 75)
        text = 'ISBN: ' + isbn
        left, top, right, bottom = draw.textbbox(position, text, font=font)
        draw.rectangle((left-5, top-5, right+5, bottom+5), fill="white")
        draw.text(position, text, fill="black", font=font)

        result.append(isbn)
    
    # save new image with isbns and bounding boxes
    img_io = io.BytesIO()
    image.save(img_io, 'JPEG')
    img_io.seek(0)
    img_io_base64 = base64.b64encode(img_io.getvalue()).decode('utf-8')
    
    return jsonify({
        'image': f"data:image/jpeg;base64,{img_io_base64}",
        'isbns': result
    })

@app.route("/api/genSpreadsheet", methods=["POST"])
def get_spreadsheet():
    isbns = request.form.get("isbns")
    
    # unstringify list
    isbns = isbns.strip(']["').split('","')

    prices_list, conditions_list = [], []
    temp = []
    for isbn in isbns:
        a, b = get_listings(isbn)
        if a == None or len(a) == 0: continue
        temp.append(isbn)

        prices_list.append(a)
        conditions_list.append(b)

    # Rebuild isbn list with only valid isbns
    isbns = temp

    # base spreadsheet
    workbook = openpyxl.load_workbook('template.xlsx')

    # if there are any valid books to be sold
    if (len(prices_list) != 0):
        prices = get_prices(prices_list, conditions_list)

        sheet = workbook.active
        length = len(prices)
        green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for i in range(length):
            cell = lambda l: l + str(i+4)
            row = [1, 0, 0, prices[i], None, isbns[i], "ISBN", "New"]

            for col, value in enumerate(row, start=1):
                sheet.cell(row=i+4, column=col).value = value

            sheet[cell("C")].fill = green
            sheet[cell("U")] = "Amazon_NA"
            sheet[cell("AB")] = "NO"
            
            for col in range(42, 47):
                sheet.cell(row=i+4, column=col).value = "not_applicable"

    io_stream = io.BytesIO()
    workbook.save(io_stream)
    io_stream.seek(0)

    return send_file(
        io_stream,
        as_attachment=True,
        download_name='processing-summary.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def get_listings(isbn_13):
    isbn_string = str(isbn_13)
    check_digit = 0
    for i in range(9):
        check_digit += (10 - i) * int(isbn_string[3 + i])

    check_digit = str(11 - (check_digit % 11))

    if check_digit == 10:
        check_digit = "X"

    isbn_10 = isbn_string[slice(3, 12)] + check_digit

    link = f"https://www.amazon.com/dp/{isbn_10}"

    driver = webdriver.Chrome()
    driver.get(link)

    # Keep retrying the Captcha until a success is found/No Captcha, Every attempt changes the captcha on amazon
    while True:
        try:
            try:
                listings = driver.find_element(By.XPATH, "//*[@id='dynamic-aod-ingress-box']/div/div[2]/a/span/span[1]")
                break
            except:
                driver.save_screenshot("./captcha.png")
                img = Image.open("./captcha.png")

                # Crop center of image to get captcha
                width, height = img.size
                left = width / 4
                top = height / 4
                right = 3 * width / 4
                bottom = 3 * height / 4
                img = img.crop((left, top, right, bottom))
                img.save('./cropped.png')

                # Read the screenshot and get the characters
                solution = reader.readtext("./cropped.png", detail=1)

                # Input the solution
                captcha_input = driver.find_element(By.XPATH, "//*[@id='captchacharacters']")
                captcha_input.send_keys(solution[0][1])

                button = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[3]/div/div/form/div[2]/div/span")
                button.click()
        except:
            return None, None, None

    if listings == None: exit()

    listings.click()

    driver.implicitly_wait(2)

    counter = 1

    # Loads the first 30 items
    while counter < 3:
        try:
            frame = driver.find_element(By.XPATH, f"//*[@id='aod-price-{counter*10}']/div/span/span[1]")
            driver.execute_script("arguments[0].scrollIntoView(true)", frame);
            time.sleep(1.5)
        except:
            break

        counter+=1

    # Loads the remaining items
    while True:
        try:
            new = driver.find_element(By.XPATH, "//*[@id='aod-show-more-offers']")
            new.click()
        except:
            break;

    count = 0
    prices_list = []

    # Get all price listings
    while True:
        try:
            price = driver.find_element(By.XPATH, f"//*[@id='aod-price-{count}']/div/span/span[1]").get_attribute("textContent")
            element = price.replace("$", "")
            element = float(element)
            prices_list.append(element)
        except:
            break
        count += 1

    quality = driver.find_elements(By.XPATH, "//*[@id='aod-offer-heading']/h5")
    conditions_list = []

    for index, qual in enumerate(quality):
        # There is one extra element that has this XPATH, duplicating the quality of the first listing, so it's ignored
        if(index == 0):
            continue

        # Removes extra characters, all used books have an actual quality to them
        element = " ".join(qual.get_attribute("textContent").split())
        element = element.replace("Used - ", "").replace("Collectible - ", "")

        # Convert quality to numerical value
        conditions_list.append(condition_stoi[element])

    driver.close()

    return prices_list, conditions_list

# Price Prediction

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
price_model = Model().to(device)
price_model.load_state_dict(torch.load(f="pricePrediction.pt"))
price_model.eval()

def get_prices(prices_list, conditions_list):
    dataset = BookDataset(prices_list, conditions_list, device=device)
    dataloader = DataLoader(dataset, batch_size=len(prices_list))
    prices, conditions, _ = next(iter(dataloader))
    
    with torch.inference_mode():
        return price_model(prices, conditions).tolist()

if __name__ == "__main__":
    print("API Ready")
    serve(app, host="127.0.0.1", port=os.environ["FLASK_PORT"])